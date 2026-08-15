"""
Data Healing acceptance demo — live AI required.

Thesis under test: the existing heal-the-function loop, with no dedicated
data-handling subsystem, can adapt code to structurally drifted input
(renamed columns, reordered columns, renamed/renested API fields) while the
ORIGINAL input keeps working.

Each scenario:
  1. write a small loader module to a temp file and import it;
  2. run it on OLD data -> must succeed without healing;
  3. run it on NEW (drifted) data -> triggers healing, which rewrites the file;
  4. re-import the healed file and assert BOTH old and new inputs produce the
     same expected business result.

Skipped automatically when no usable AI provider is configured.

Run with:  python -m pytest tests/test_data_drift.py -v
"""

import importlib.util
import sys
from pathlib import Path

import pytest


def _ai_ready() -> bool:
    """True when the user config points at a real (non-placeholder) provider."""
    cfg_path = Path.home() / ".healing_agent" / "healing_agent_config.py"
    if not cfg_path.exists():
        return False
    try:
        spec = importlib.util.spec_from_file_location("_drift_cfg", cfg_path)
        cfg = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(cfg)
        provider = getattr(cfg, "AI_PROVIDER", "")
        block = getattr(cfg, provider.upper(), {}) if provider else {}
        key = str(block.get("api_key") or "")
        return len(key) > 10 and "XXX" not in key and "your-" not in key
    except Exception:
        return False


pytestmark = pytest.mark.skipif(
    not _ai_ready(), reason="No usable AI provider configured (live demo test)"
)


# --- Scenario fixtures -------------------------------------------------------
# Same business data in every old/new pair, so the expected result is identical.

CSV_LOADER = '''
import healing_agent

@healing_agent
def load_sales(csv_text):
    import csv, io
    rows = list(csv.DictReader(io.StringIO(csv_text)))
    total = 0
    customers = []
    for r in rows:
        total += int(r["amount"])
        customers.append(r["customer"])
    return {"total": total, "customers": sorted(customers)}
'''

OLD_CSV = "date,customer,amount\n2026-01-05,Alfa Kft,1200\n2026-01-06,Beta Zrt,800\n"
# Drift: headers renamed to Hungarian.
NEW_CSV_RENAMED = "datum,ugyfel,osszeg\n2026-01-05,Alfa Kft,1200\n2026-01-06,Beta Zrt,800\n"

INDEX_LOADER = '''
import healing_agent

@healing_agent
def load_sales(csv_text):
    import csv, io
    reader = csv.reader(io.StringIO(csv_text))
    next(reader)  # skip header
    total = 0
    customers = []
    for row in reader:
        total += int(row[2])
        customers.append(row[1])
    return {"total": total, "customers": sorted(customers)}
'''

# Drift: column ORDER changed (amount is no longer index 2).
NEW_CSV_REORDERED = "customer,amount,date\nAlfa Kft,1200,2026-01-05\nBeta Zrt,800,2026-01-06\n"

API_LOADER = '''
import healing_agent

@healing_agent
def summarize_orders(payload):
    items = payload["data"]["items"]
    total = 0
    names = []
    for item in items:
        total += int(item["price"])
        names.append(item["name"])
    return {"total": total, "names": sorted(names)}
'''

OLD_PAYLOAD = {
    "data": {"items": [{"name": "Alfa Kft", "price": 1200}, {"name": "Beta Zrt", "price": 800}]}
}
# Drift: the API response was renamed and renested.
NEW_PAYLOAD = {
    "result": {"records": [{"title": "Alfa Kft", "amount": 1200}, {"title": "Beta Zrt", "amount": 800}]}
}

EXPECTED = {"total": 2000, "customers": ["Alfa Kft", "Beta Zrt"]}
EXPECTED_API = {"total": 2000, "names": ["Alfa Kft", "Beta Zrt"]}

# --- Harder scenarios --------------------------------------------------------

DATE_LOADER = '''
import healing_agent

@healing_agent
def monthly_totals(csv_text):
    import csv, io
    from datetime import datetime
    rows = list(csv.DictReader(io.StringIO(csv_text)))
    totals = {}
    for r in rows:
        d = datetime.strptime(r["date"], "%Y-%m-%d")
        key = f"{d.year}-{d.month:02d}"
        totals[key] = totals.get(key, 0) + int(r["amount"])
    return totals
'''

OLD_CSV_DATES = "date,customer,amount\n2026-01-15,Alfa Kft,1200\n2026-02-16,Beta Zrt,800\n"
# Drift: dates switch to Hungarian DD.MM.YYYY (days > 12, so unambiguous).
NEW_CSV_DATES = "date,customer,amount\n15.01.2026,Alfa Kft,1200\n16.02.2026,Beta Zrt,800\n"
EXPECTED_DATES = {"2026-01": 1200, "2026-02": 800}

# The exception surfaces inside an UNdecorated helper; only the decorated
# entry point can be rewritten. The fix must adapt at the boundary.
HELPER_LOADER = '''
import healing_agent

def parse_row(row):
    return {"customer": row["customer"], "amount": int(row["amount"])}

@healing_agent
def load_sales(csv_text):
    import csv, io
    rows = list(csv.DictReader(io.StringIO(csv_text)))
    total = 0
    customers = []
    for r in rows:
        parsed = parse_row(r)
        total += parsed["amount"]
        customers.append(parsed["customer"])
    return {"total": total, "customers": sorted(customers)}
'''

SCENARIOS = [
    ("csv_renamed_headers", CSV_LOADER, "load_sales", OLD_CSV, NEW_CSV_RENAMED, EXPECTED),
    ("csv_reordered_columns", INDEX_LOADER, "load_sales", OLD_CSV, NEW_CSV_REORDERED, EXPECTED),
    ("api_renamed_renested", API_LOADER, "summarize_orders", OLD_PAYLOAD, NEW_PAYLOAD, EXPECTED_API),
    ("date_format_drift", DATE_LOADER, "monthly_totals", OLD_CSV_DATES, NEW_CSV_DATES, EXPECTED_DATES),
    ("error_in_helper_function", HELPER_LOADER, "load_sales", OLD_CSV, NEW_CSV_RENAMED, EXPECTED),
]


# --- Harness -----------------------------------------------------------------

def _import_module(path: Path, name: str):
    spec = importlib.util.spec_from_file_location(name, path)
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


@pytest.mark.parametrize(
    "name,source,func_name,old_data,new_data,expected",
    SCENARIOS,
    ids=[s[0] for s in SCENARIOS],
)
def test_drifted_input_is_healed(tmp_path, name, source, func_name, old_data, new_data, expected):
    module_path = tmp_path / f"loader_{name}.py"
    module_path.write_text(source, encoding="utf-8")
    module_name = f"drift_demo_{name}"

    # 1) OLD data works out of the box (no healing involved).
    module = _import_module(module_path, module_name)
    assert getattr(module, func_name)(old_data) == expected

    # 2) NEW drifted data triggers healing; the wrapper should return the
    #    repaired result already.
    result = getattr(module, func_name)(new_data)
    assert result == expected, f"healed call on drifted data returned {result!r}"

    # 3) The healed source must handle BOTH formats: re-import fresh and
    #    verify old AND new inputs. This is the actual data-healing claim.
    healed = _import_module(module_path, module_name + "_healed")
    assert getattr(healed, func_name)(old_data) == expected, "old format broke after healing"
    assert getattr(healed, func_name)(new_data) == expected, "new format not handled after healing"


# --- Guardrail: unmappable drift must FAIL, not fabricate --------------------

MISSING_COLUMN_LOADER = '''
import healing_agent

@healing_agent(MAX_ATTEMPTS=1)
def load_sales(csv_text):
    import csv, io
    rows = list(csv.DictReader(io.StringIO(csv_text)))
    total = 0
    customers = []
    for r in rows:
        total += int(r["amount"])
        customers.append(r["customer"])
    return {"total": total, "customers": sorted(customers)}
'''

# The amount column is GONE, not renamed. No honest mapping exists.
NEW_CSV_NO_AMOUNT = "datum,ugyfel\n2026-01-15,Alfa Kft\n2026-02-16,Beta Zrt\n"


def test_unmappable_drift_raises_instead_of_fabricating(tmp_path):
    """Required business data is missing entirely: healing must surface an
    error, never invent amounts. Old-format inputs must keep working on the
    (possibly rewritten) source afterwards."""
    module_path = tmp_path / "loader_missing_column.py"
    module_path.write_text(MISSING_COLUMN_LOADER, encoding="utf-8")
    module = _import_module(module_path, "drift_demo_missing_column")

    # Old data works.
    assert module.load_sales(OLD_CSV) == EXPECTED

    # Unmappable data must end in an exception, not a fabricated result.
    with pytest.raises(Exception):
        module.load_sales(NEW_CSV_NO_AMOUNT)

    # Whatever healing did to the file, the old format must still work…
    healed = _import_module(module_path, "drift_demo_missing_column_healed")
    assert healed.load_sales(OLD_CSV) == EXPECTED, "old format broke after healing attempt"


# Adversarial variant: the amount column is gone, but a DECOY numeric column
# (order number) is present. A lazy fix would sum order numbers as amounts.
NEW_CSV_DECOY_NUMERIC = (
    "datum,ugyfel,rendeles_szam\n"
    "2026-01-15,Alfa Kft,5001\n"
    "2026-02-16,Beta Zrt,5002\n"
)


def test_decoy_numeric_column_is_not_mistaken_for_amount(tmp_path):
    """Missing required column + an unrelated numeric column: the healed code
    must not silently repurpose the decoy as business data."""
    module_path = tmp_path / "loader_decoy_column.py"
    module_path.write_text(MISSING_COLUMN_LOADER, encoding="utf-8")
    module = _import_module(module_path, "drift_demo_decoy_column")

    assert module.load_sales(OLD_CSV) == EXPECTED

    with pytest.raises(Exception):
        module.load_sales(NEW_CSV_DECOY_NUMERIC)

    healed = _import_module(module_path, "drift_demo_decoy_column_healed")
    assert healed.load_sales(OLD_CSV) == EXPECTED, "old format broke after healing attempt"


# --- Excel workbook drift ------------------------------------------------------
# Three drift layers at once: the sheet was renamed, title/empty rows appeared
# above the header row, and the headers themselves were translated.
#
# Unlike the CSV scenarios, the function argument is only a PATH: the model
# cannot see the drifted content up front. Each repair round discovers more
# (sheet names -> title row -> actual headers) through its own error messages,
# so the attempt budget must scale with the number of drift layers.

EXCEL_LOADER = '''
import healing_agent

@healing_agent(MAX_ATTEMPTS=5)
def load_sales(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Sales"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    total = 0
    customers = []
    for row in rows[1:]:
        rec = dict(zip(headers, row))
        total += int(rec["amount"])
        customers.append(rec["customer"])
    return {"total": total, "customers": sorted(customers)}
'''


def _write_workbook(path, sheet_name, prefix_rows, headers, data_rows):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in prefix_rows:
        ws.append(row)
    ws.append(headers)
    for row in data_rows:
        ws.append(row)
    wb.save(path)


def test_excel_workbook_drift(tmp_path):
    pytest.importorskip("openpyxl")

    old_xlsx = tmp_path / "sales_old.xlsx"
    new_xlsx = tmp_path / "sales_new.xlsx"
    _write_workbook(
        old_xlsx, "Sales", [],
        ["date", "customer", "amount"],
        [["2026-01-15", "Alfa Kft", 1200], ["2026-02-16", "Beta Zrt", 800]],
    )
    _write_workbook(
        new_xlsx, "Ertekesites 2026",
        [["Havi ertekesitesi riport", None, None], [None, None, None]],
        ["datum", "ugyfel", "osszeg"],
        [["2026-01-15", "Alfa Kft", 1200], ["2026-02-16", "Beta Zrt", 800]],
    )

    module_path = tmp_path / "loader_excel_drift.py"
    module_path.write_text(EXCEL_LOADER, encoding="utf-8")
    module = _import_module(module_path, "drift_demo_excel")

    assert module.load_sales(str(old_xlsx)) == EXPECTED

    result = module.load_sales(str(new_xlsx))
    assert result == EXPECTED, f"healed call on drifted workbook returned {result!r}"

    healed = _import_module(module_path, "drift_demo_excel_healed")
    assert healed.load_sales(str(old_xlsx)) == EXPECTED, "old workbook broke after healing"
    assert healed.load_sales(str(new_xlsx)) == EXPECTED, "new workbook not handled after healing"


# --- Quarantine semantics -------------------------------------------------------
# The original contract already quarantines malformed rows instead of failing
# the whole batch. Healing the header drift must PRESERVE that semantics:
# unmappable/invalid rows stay rejected, and are never fabricated into results.

QUARANTINE_LOADER = '''
import healing_agent

@healing_agent
def load_sales(csv_text):
    import csv, io
    rows = list(csv.DictReader(io.StringIO(csv_text)))
    total = 0
    customers = []
    rejected = 0
    for r in rows:
        try:
            total += int(r["amount"])
            customers.append(r["customer"])
        except (ValueError, TypeError):
            rejected += 1  # quarantine malformed rows, keep processing
    return {"total": total, "customers": sorted(customers), "rejected": rejected}
'''

# Renamed headers AND two malformed rows (non-numeric / empty amount).
# Distinct customer names on the bad rows: if healing fabricates amounts
# (e.g. treats N/A as 0), those names would appear and the assertion fails.
NEW_CSV_MIXED = (
    "datum,ugyfel,osszeg\n"
    "2026-01-15,Alfa Kft,1200\n"
    "2026-01-20,Gamma Bt,N/A\n"
    "2026-02-16,Beta Zrt,800\n"
    "2026-02-20,Delta Kkt,\n"
)

EXPECTED_Q_OLD = {"total": 2000, "customers": ["Alfa Kft", "Beta Zrt"], "rejected": 0}
EXPECTED_Q_NEW = {"total": 2000, "customers": ["Alfa Kft", "Beta Zrt"], "rejected": 2}


def test_mixed_records_are_quarantined_not_fabricated(tmp_path):
    module_path = tmp_path / "loader_quarantine.py"
    module_path.write_text(QUARANTINE_LOADER, encoding="utf-8")
    module = _import_module(module_path, "drift_demo_quarantine")

    # Old format, all rows valid.
    assert module.load_sales(OLD_CSV) == EXPECTED_Q_OLD

    # Drifted headers + malformed rows: healing must map the headers while
    # keeping quarantine semantics for the two bad rows.
    result = module.load_sales(NEW_CSV_MIXED)
    assert result == EXPECTED_Q_NEW, f"healed call returned {result!r}"

    healed = _import_module(module_path, "drift_demo_quarantine_healed")
    assert healed.load_sales(OLD_CSV) == EXPECTED_Q_OLD, "old format broke after healing"
    assert healed.load_sales(NEW_CSV_MIXED) == EXPECTED_Q_NEW, "quarantine semantics lost after healing"
